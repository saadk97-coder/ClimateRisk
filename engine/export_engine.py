"""
XLSX export engine with assurance-oriented metadata, lineage, and methodology notes.
"""

from __future__ import annotations

import io
from typing import Optional

import numpy as np
import pandas as pd

from engine.data_sources import DATA_SOURCE_REGISTRY
from engine.governance import (
    BASELINE_METHOD,
    DCF_POSITIONING,
    MODEL_SCOPE,
    PLATFORM_NAME,
    RESULTS_POSITIONING,
    runtime_metadata,
)
from engine.scenario_model import HAZARD_SCALING_SOURCES, SCENARIOS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

HDR_FILL = "1F4E79"
SUB_FILL = "2E75B6"

VULNERABILITY_ROWS = [
    {
        "Component": "Vulnerability",
        "Item": "Flood",
        "Citation": "FEMA HAZUS 6.0; JRC Global Flood DDFs (Huizinga et al. 2017)",
        "URL": "https://www.fema.gov/flood-maps/products-tools/hazus",
        "Notes": "Depth-damage functions mapped to asset-type curve keys.",
    },
    {
        "Component": "Vulnerability",
        "Item": "Wind",
        "Citation": "FEMA HAZUS MH Hurricane Technical Manual; IBHS FORTIFIED",
        "URL": "https://www.fema.gov/sites/default/files/2020-09/fema_hazus-hurricane-technical-manual.pdf",
        "Notes": "Wind fragility curves applied as asset-type specific vulnerability functions.",
    },
    {
        "Component": "Vulnerability",
        "Item": "Wildfire",
        "Citation": "Syphard et al. (2012); FEMA HAZUS Wildfire",
        "URL": "https://doi.org/10.1890/ES12-00197.1",
        "Notes": "Flame-length curves are screening-level structural damage proxies.",
    },
    {
        "Component": "Vulnerability",
        "Item": "Heat",
        "Citation": "IEA (2018); ILO (2019); Zhao et al. (2021)",
        "URL": "https://www.iea.org/reports/the-future-of-cooling",
        "Notes": "Heat curves combine cooling-cost escalation and productivity-loss proxies.",
    },
    {
        "Component": "Vulnerability",
        "Item": "Coastal Flood",
        "Citation": "Vousdoukas et al. (2018); Muis et al. (2020)",
        "URL": "https://doi.org/10.1038/s41467-018-04692-w",
        "Notes": "Coastal curves use storm-surge depth damage relationships.",
    },
]


def _hdr_style(ws, row: int, fill_hex: str = HDR_FILL):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF" if fill_hex in (HDR_FILL, SUB_FILL) else "000000")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_width: int = 10, max_width: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=min_width)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_width), max_width)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, header_fill: str = HDR_FILL):
    if df is None or df.empty:
        ws.cell(start_row, 1, "No data")
        return
    headers = list(df.columns)
    for j, header in enumerate(headers, 1):
        ws.cell(start_row, j, header)
    _hdr_style(ws, start_row, header_fill)
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val if not (isinstance(val, float) and np.isnan(val)) else None)
    _autofit(ws)


def _ordered_metadata(metadata: Optional[dict]) -> dict:
    merged = runtime_metadata()
    if metadata:
        merged.update({k: v for k, v in metadata.items() if v not in (None, "")})
    return merged


def _add_cover(ws, metadata: Optional[dict]) -> int:
    ws.cell(1, 1, PLATFORM_NAME)
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, MODEL_SCOPE)
    ws.cell(3, 1, RESULTS_POSITIONING)
    ws.cell(4, 1, BASELINE_METHOD)
    row = 6
    for key, value in _ordered_metadata(metadata).items():
        ws.cell(row, 1, key)
        ws.cell(row, 2, str(value))
        row += 1
    return row + 1


def _metadata_df(metadata: Optional[dict]) -> pd.DataFrame:
    return pd.DataFrame(
        [{"Field": key, "Value": value} for key, value in _ordered_metadata(metadata).items()]
    )


def _source_rows(annual_damages_df: Optional[pd.DataFrame], scenarios: Optional[list], override_records: Optional[list[dict]]) -> pd.DataFrame:
    rows: list[dict] = []

    for scenario_id in scenarios or []:
        info = SCENARIOS.get(scenario_id, {})
        rows.append(
            {
                "Component": "Scenario set",
                "Item": info.get("label", scenario_id),
                "Citation": f"{info.get('provider', 'Scenario provider')} | {info.get('ssp', '')}",
                "URL": info.get("source_url", ""),
                "Notes": info.get("description", ""),
            }
        )

    hazards_used = []
    if annual_damages_df is not None and not annual_damages_df.empty and "hazard" in annual_damages_df.columns:
        hazards_used = sorted(set(annual_damages_df["hazard"].dropna().astype(str)))
    for hazard in hazards_used:
        src = HAZARD_SCALING_SOURCES.get(hazard, {})
        if src:
            rows.append(
                {
                    "Component": "Hazard scaling",
                    "Item": hazard,
                    "Citation": src.get("citation", ""),
                    "URL": src.get("url", ""),
                    "Notes": "Forward-looking change is applied through scenario multipliers.",
                }
            )

    source_keys = []
    if annual_damages_df is not None and not annual_damages_df.empty and "data_source" in annual_damages_df.columns:
        source_keys = sorted(set(annual_damages_df["data_source"].dropna().astype(str)))
    for source_key in source_keys:
        info = DATA_SOURCE_REGISTRY.get(source_key, {})
        rows.append(
            {
                "Component": "Hazard data",
                "Item": info.get("name", source_key),
                "Citation": info.get("citation", ""),
                "URL": info.get("url", ""),
                "Notes": info.get("description", ""),
            }
        )

    rows.extend(VULNERABILITY_ROWS)

    if override_records:
        rows.append(
            {
                "Component": "Manual overrides",
                "Item": "Analyst-entered hazard overrides",
                "Citation": "Session-state manual override records",
                "URL": "",
                "Notes": (
                    "Overrides replace fetched baseline intensities for specified asset-hazard pairs. "
                    "Provenance is documented on the Manual Overrides sheet."
                ),
            }
        )

    return pd.DataFrame(rows)


def _method_notes_df(annual_damages_df: Optional[pd.DataFrame], override_records: Optional[list[dict]]) -> pd.DataFrame:
    hazards_used = []
    if annual_damages_df is not None and not annual_damages_df.empty and "hazard" in annual_damages_df.columns:
        hazards_used = sorted(set(annual_damages_df["hazard"].dropna().astype(str)))

    rows = [
        {"Topic": "Scope", "Detail": RESULTS_POSITIONING},
        {"Topic": "Baseline method", "Detail": BASELINE_METHOD},
        {
            "Topic": "Acute hazard EAD",
            "Detail": (
                "Acute hazards use trapezoidal integration of loss over annual exceedance probability "
                "across the discrete return-period curve."
            ),
        },
    ]

    if "water_stress" in hazards_used:
        rows.append(
            {
                "Topic": "Water stress method",
                "Detail": (
                    "Water stress is treated as a chronic hazard. Expected annual damage is computed as "
                    "the RP50 damage fraction multiplied by replacement value, not by EP-curve integration."
                ),
            }
        )
    if "flood" in hazards_used:
        rows.append(
            {
                "Topic": "Flood limitation",
                "Detail": (
                    "Flood depths are screening-level proxies derived from gridded climate and hydrological "
                    "data. They are not local hydraulic depth simulations."
                ),
            }
        )
    if "coastal_flood" in hazards_used:
        rows.append(
            {
                "Topic": "Coastal flood method",
                "Detail": (
                    "Coastal flood combines multiplicative scenario scaling with additive sea-level rise and "
                    "asset freeboard adjustments."
                ),
            }
        )
    if override_records:
        rows.append(
            {
                "Topic": "Manual overrides",
                "Detail": (
                    "Manual overrides are analyst-entered intensities that supersede fetched hazard data for "
                    "specific asset-hazard pairs. They should be reviewed with the recorded evidence source."
                ),
            }
        )

    return pd.DataFrame(rows)


def export_results_xlsx(
    asset_results_df: pd.DataFrame,
    annual_damages_df: Optional[pd.DataFrame],
    portfolio_summary: Optional[dict],
    scenarios: list,
    metadata: dict,
    override_records: Optional[list[dict]] = None,
) -> bytes:
    if not _HAS_OPENPYXL:
        buf = io.BytesIO()
        asset_results_df.to_excel(buf, index=False)
        return buf.getvalue()

    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Portfolio Summary"
    next_row = _add_cover(ws_sum, metadata)
    if portfolio_summary:
        _write_df(ws_sum, pd.DataFrame([{"Metric": k, "Value": v} for k, v in portfolio_summary.items()]), start_row=next_row)
    _autofit(ws_sum)

    ws_meta = wb.create_sheet("Run Metadata")
    _write_df(ws_meta, _metadata_df(metadata))

    ws_ar = wb.create_sheet("Asset Results")
    _write_df(ws_ar, asset_results_df)

    if annual_damages_df is not None and not annual_damages_df.empty:
        ws_ann = wb.create_sheet("Annual Damages 2025-2050")
        _write_df(ws_ann, annual_damages_df)

        try:
            pivot = annual_damages_df.pivot_table(
                index="year", columns="scenario_id", values="ead", aggfunc="sum"
            ).reset_index()
            ws_piv = wb.create_sheet("Annual EAD Pivot")
            _write_df(ws_piv, pivot)
        except Exception:
            pass

        ws_sc = wb.create_sheet("Scenario Comparison")
        sc_rows = []
        for scenario_id in scenarios:
            sc_df = annual_damages_df[annual_damages_df["scenario_id"] == scenario_id]
            if sc_df.empty:
                continue
            sc_label = SCENARIOS.get(scenario_id, {}).get("label", scenario_id)
            currency = metadata.get("currency_symbol", "")
            sc_rows.append(
                {
                    "Scenario": sc_label,
                    "Scenario ID": scenario_id,
                    f"Total EAD 2025-2050 ({currency})": round(float(sc_df["ead"].sum()), 2),
                    f"Total PV Damages ({currency})": round(float(sc_df["pv"].sum()), 2),
                    f"Mean Annual EAD ({currency})": round(float(sc_df.groupby("year")["ead"].sum().mean()), 2),
                }
            )
        _write_df(ws_sc, pd.DataFrame(sc_rows))

    ws_src = wb.create_sheet("Sources & Methodology")
    _write_df(ws_src, _source_rows(annual_damages_df, scenarios, override_records))

    ws_method = wb.create_sheet("Method Notes")
    _write_df(ws_method, _method_notes_df(annual_damages_df, override_records))

    if override_records:
        ws_ov = wb.create_sheet("Manual Overrides")
        _write_df(ws_ov, pd.DataFrame(override_records))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_audit_xlsx(
    audit_df: pd.DataFrame,
    metadata: dict,
    override_records: Optional[list[dict]] = None,
) -> bytes:
    if not _HAS_OPENPYXL:
        buf = io.BytesIO()
        audit_df.to_excel(buf, index=False)
        return buf.getvalue()

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation Audit"
    next_row = _add_cover(ws, metadata)
    ws.cell(next_row, 1, "Step-by-step calculation trace")
    ws.cell(next_row, 1).font = Font(bold=True, italic=True)
    _write_df(ws, audit_df, start_row=next_row + 2)

    ws_meta = wb.create_sheet("Run Metadata")
    _write_df(ws_meta, _metadata_df(metadata))

    ws_method = wb.create_sheet("Method Notes")
    _write_df(
        ws_method,
        pd.DataFrame(
            [
                {"Topic": "Scope", "Detail": RESULTS_POSITIONING},
                {"Topic": "Baseline method", "Detail": BASELINE_METHOD},
                {
                    "Topic": "Audit purpose",
                    "Detail": "This workbook is a calculation trace for the selected asset, scenario, year, and hazard.",
                },
            ]
        ),
    )

    if override_records:
        ws_ov = wb.create_sheet("Manual Overrides")
        _write_df(ws_ov, pd.DataFrame(override_records))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_adaptation_xlsx(adaptation_df: pd.DataFrame, frontier_df: pd.DataFrame) -> bytes:
    if not _HAS_OPENPYXL:
        buf = io.BytesIO()
        adaptation_df.to_excel(buf, index=False)
        return buf.getvalue()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Adaptation CBA"
    _write_df(ws1, adaptation_df)

    ws2 = wb.create_sheet("Portfolio Frontier")
    _write_df(ws2, frontier_df)

    ws3 = wb.create_sheet("Methodology")
    notes = [
        ["Metric", "Formula", "Source"],
        ["Capex", "asset_value x capex_pct/100", "FEMA BCA Guide; EA FCERM"],
        ["Annual Opex", "capex x opex_annual_pct/100", "FEMA BCA Guide"],
        ["NPV Benefits", "Sum of avoided_EAD discounted over design life", "Standard NPV"],
        ["Cost-Benefit Ratio", "NPV Benefits / (capex + NPV opex)", "FEMA BCA; HM Treasury Green Book"],
        ["Payback Period", "capex / avoided_EAD_annual", "Standard payback formula"],
        ["Avoided EAD", "baseline_EAD x damage_reduction_pct/100", "Per-measure effectiveness"],
        ["Scope", RESULTS_POSITIONING, ""],
    ]
    _write_df(ws3, pd.DataFrame(notes[1:], columns=notes[0]))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_dcf_xlsx(dcf_results: list, scenario_comparison_df: pd.DataFrame) -> bytes:
    if not _HAS_OPENPYXL:
        buf = io.BytesIO()
        scenario_comparison_df.to_excel(buf, index=False)
        return buf.getvalue()

    wb = Workbook()

    ws_sc = wb.active
    ws_sc.title = "Scenario Comparison"
    _write_df(ws_sc, scenario_comparison_df)

    for result in dcf_results:
        safe_name = result.label[:28].replace("/", "-")
        ws = wb.create_sheet(safe_name)
        _write_df(ws, result.annual_detail)

    ws_meta = wb.create_sheet("Run Metadata")
    _write_df(
        ws_meta,
        pd.DataFrame(
            [
                {"Field": "Platform", "Value": PLATFORM_NAME},
                {"Field": "Scope", "Value": DCF_POSITIONING},
                {"Field": "Baseline method", "Value": BASELINE_METHOD},
            ]
        ),
    )

    ws_m = wb.create_sheet("Methodology")
    meth = [
        ["Item", "Description", "Reference"],
        ["Base NPV", "NPV = discounted cash flows plus discounted terminal value", "Standard DCF"],
        ["Climate-Adjusted NPV", "Base cash flows less annual climate damages, discounted at WACC", "BSR Climate Strategy Framework"],
        ["Climate Risk Premium", "Optional uplift applied to WACC as a sensitivity input", "TCFD / BSR"],
        ["Terminal Value", "Gordon Growth approximation", "Standard finance"],
        ["Scenario Weighting", "Probability-weighted average of scenario NPVs", "TCFD scenario analysis"],
        ["Scope", DCF_POSITIONING, ""],
    ]
    _write_df(ws_m, pd.DataFrame(meth[1:], columns=meth[0]))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def df_to_xlsx(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    buf = io.BytesIO()
    if _HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        _write_df(ws, df)
        wb.save(buf)
    else:
        df.to_excel(buf, index=False)
    return buf.getvalue()
